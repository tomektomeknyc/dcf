# excel_bridge.py
from __future__ import annotations
from pathlib import Path
import yaml
from openpyxl import load_workbook
import pandas as pd

# ---- config ----
REQUIRED_SCALARS = [
    "BR_VERSION",
    "BR_TICKER",
    "BR_CURRENCY",
    "BR_WACC",
    "BR_RF",     
    "BR_RD",
    "BR_TAX",     
    "BR_SHARES",
    "BR_G_LT",
]

REQUIRED_RANGES  = ["BR_FCFF_hist"]  # Year|FCFF (two columns, header row included)

class BridgeData(pd.NamedAgg):  # just a tag; we'll return dicts/DataFrames
    pass

def load_sources(yaml_path: str | Path = "sources.yaml") -> dict:
    with open(yaml_path, "r") as f:
        cfg = yaml.safe_load(f) or {}
    return cfg.get("tickers", {})

def open_book(xlsx_path: str | Path):
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    return load_workbook(path, data_only=True, read_only=True)

def _get_scalar(wb, name: str):
    if name not in wb.defined_names:
        raise KeyError(f"Missing named cell: {name}")
    dest = list(wb.defined_names[name].destinations)[0]  # (sheet, ref)
    ws = wb[dest[0]]
    return ws[dest[1]].value

def _get_table(wb, name: str) -> pd.DataFrame:
    if name not in wb.defined_names:
        raise KeyError(f"Missing named range: {name}")
    sheet, ref = list(wb.defined_names[name].destinations)[0]
    ws = wb[sheet]
    rows = ws[ref]
    data = [[cell.value for cell in row] for row in rows]
    # assume first row is headers
    return pd.DataFrame(data[1:], columns=[str(h) for h in data[0]])

def load_bridge_for(ticker: str, sources: dict) -> dict:
    meta = sources.get(ticker)
    if not meta:
        raise KeyError(f"No source configured for ticker key: {ticker}")

    wb = open_book(meta["path"])

    # Scalars
    scalars = {k: _get_scalar(wb, k) for k in REQUIRED_SCALARS}

    # Ranges/tables
    tables = {k: _get_table(wb, k) for k in REQUIRED_RANGES}

    # Basic sanity
    if str(scalars["BR_VERSION"]).strip() not in {"1", "1.0", "v1", "V1"}:
        raise ValueError("Bridge version not recognized (want 1.0).")

    return {
        "scalars": scalars,
        "tables": tables,
        "path": meta["path"],
        "type": meta.get("type", "presentation"),
    }
